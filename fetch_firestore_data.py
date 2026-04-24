"""
Firestore to Excel Exporter
Fetches quiz results from Firebase Firestore and exports to Excel

Setup:
1. Install dependencies:
   pip install firebase-admin pandas openpyxl

2. Go to Firebase Console > Project Settings > Service Accounts
3. Click "Generate new private key"
4. Save the JSON file as 'serviceAccountKey.json' in the same folder

Run:
python fetch_firestore_data.py
"""

import firebase_admin
from firebase_admin import credentials, firestore
import pandas as pd
from datetime import datetime
import os


def initialize_firebase():
    """Initialize Firebase Admin SDK"""
    cred_path = 'serviceAccountKey.json'
    
    if not os.path.exists(cred_path):
        print(f"❌ Error: '{cred_path}' not found!")
        print("\n📋 Setup Instructions:")
        print("1. Go to Firebase Console > Project Settings > Service Accounts")
        print("2. Click 'Generate new private key'")
        print("3. Save the JSON file as 'serviceAccountKey.json' in this folder")
        return None
    
    if firebase_admin._apps:
        return firestore.client()
    
    cred = credentials.Certificate(cred_path)
    firebase_admin.initialize_app(cred)
    return firestore.client()


def fetch_quizResults(db):
    """Fetch all documents from quizResults collection"""
    collection_name = 'quizResults'
    
    docs = db.collection(collection_name).stream()
    
    results = []
    for doc in docs:
        data = doc.to_dict()
        data['document_id'] = doc.id
        
        # Flatten studentInfo object
        if 'studentInfo' in data and data['studentInfo']:
            student_info = data['studentInfo']
            data['firstName'] = student_info.get('firstName', '')
            data['lastName'] = student_info.get('lastName', '')
            data['rollNumber'] = student_info.get('rollNumber', '')
            del data['studentInfo']  # Remove nested object
        
        # Flatten results object
        if 'results' in data and data['results']:
            results_data = data['results']
            data['totalMarks'] = results_data.get('totalMarks', 0)
            data['score'] = results_data.get('totalEarned', 0)
            data['percentage'] = results_data.get('percentage', 0)
            data['grade'] = results_data.get('grade', '')
            data['correctCount'] = results_data.get('correctCount', 0)
            data['wrongCount'] = results_data.get('wrongCount', 0)
            data['skippedCount'] = results_data.get('skippedCount', 0)
            del data['results']  # Remove nested object
        
        if 'timestamp' in data and data['timestamp']:
            data['timestamp'] = data['timestamp'].strftime('%Y-%m-%d %H:%M:%S')
        
        results.append(data)
    
    return results


def export_to_excel(data, filename=None):
    """Export data to Excel with formatting"""
    if not data:
        print("⚠️  No data to export!")
        return
    
    if filename is None:
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        filename = f'quizResults_{timestamp}.xlsx'
    
    df = pd.DataFrame(data)
    
    column_order = [
        'document_id',
        'timestamp',
        'firstName',
        'lastName',
        'rollNumber',
        'class',
        'subject',
        'examName',
        'teacher',
        'invigilator',
        'totalMarks',
        'score',
        'percentage',
        'grade',
        'correctCount',
        'wrongCount',
        'skippedCount',
        'totalQuestions',
        'timeLimit'
    ]
    
    existing_cols = [col for col in column_order if col in df.columns]
    remaining_cols = [col for col in df.columns if col not in column_order]
    df = df[existing_cols + remaining_cols]
    
    df = df.rename(columns={
        'document_id': 'ID',
        'timestamp': 'Date & Time',
        'firstName': 'First Name',
        'lastName': 'Last Name',
        'rollNumber': 'Roll No.',
        'class': 'Class',
        'subject': 'Subject',
        'examName': 'Exam',
        'teacher': 'Teacher',
        'invigilator': 'Invigilator',
        'totalMarks': 'Total Marks',
        'score': 'Score',
        'percentage': 'Percentage',
        'grade': 'Grade',
        'correctCount': 'Correct',
        'wrongCount': 'Wrong',
        'skippedCount': 'Skipped',
        'totalQuestions': 'Total Questions',
        'timeLimit': 'Time Limit (min)'
    })
    
    df = df.sort_values('Date & Time', ascending=False)
    
    output_path = os.path.join('exports', filename)
    os.makedirs('exports', exist_ok=True)
    
    with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
        df.to_excel(writer, sheet_name='Results', index=False)
        
        workbook = writer.book
        worksheet = writer.sheets['Results']
        
        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
        
        header_fill = PatternFill(start_color='667EEA', end_color='764BA2', fill_type='solid')
        header_font = Font(bold=True, color='FFFFFF', size=11)
        
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        for cell in worksheet[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = thin_border
        
        column_widths = {
            'A': 15, 'B': 20, 'C': 12, 'D': 20, 'E': 10,
            'F': 15, 'G': 25, 'H': 18, 'I': 18, 'J': 12,
            'K': 10, 'L': 10, 'M': 12, 'N': 8,
            'O': 10, 'P': 10, 'Q': 15, 'R': 15
        }
        
        for col, width in column_widths.items():
            if col in worksheet.column_dimensions:
                worksheet.column_dimensions[col].width = width
        
        for row in worksheet.iter_rows(min_row=2, max_row=worksheet.max_row):
            for cell in row:
                cell.alignment = Alignment(horizontal='center', vertical='center')
                cell.border = thin_border
        
        worksheet.row_dimensions[1].height = 25
        
        from openpyxl.formatting.rule import FormulaRule
        
        # Grade color mapping (green=best, red=worst)
        grade_colors = {
            'A1': 'D4EDDA',  # Excellent - green
            'A2': 'D4EDDA',  # Excellent - green
            'B1': 'B8E6B8',  # Very Good - light green
            'B2': 'FFF8DC',  # Good - cream
            'C1': 'FFE4B5',  # Average - moccasin
            'C2': 'FFDAB9',  # Below Average - peach
            'D': 'FFB6C1',  # Poor - light pink
            'E': 'FF6B6B',  # Fail - red
        }
        
        # Find grade column index dynamically
        grade_col_idx = None
        for idx, col_name in enumerate(column_order, start=1):
            if col_name == 'grade':
                grade_col_idx = idx
                break
        
        if grade_col_idx is None:
            grade_col_idx = 13  # Fallback
        
        # Apply grade colors
        for row in worksheet.iter_rows(min_row=2, max_row=worksheet.max_row, min_col=grade_col_idx, max_col=grade_col_idx):
            for cell in row:
                grade = str(cell.value).strip() if cell.value else ''
                if grade in grade_colors:
                    cell.fill = PatternFill(start_color=grade_colors[grade], end_color=grade_colors[grade], fill_type='solid')
                    cell.font = Font(bold=True)
        
        worksheet.freeze_panes = 'A2'
    
    return output_path


def print_statistics(data):
    """Print summary statistics"""
    if not data:
        return
    
    df = pd.DataFrame(data)
    
    print("\n" + "=" * 50)
    print("📊 QUIZ RESULTS SUMMARY")
    print("=" * 50)
    
    total_students = len(df)
    print(f"📝 Total Submissions: {total_students}")
    
    if 'percentage' in df.columns:
        avg_percentage = df['percentage'].mean()
        print(f"📈 Average Score: {avg_percentage:.2f}%")
        print(f"📈 Highest Score: {df['percentage'].max():.2f}%")
        print(f"📈 Lowest Score: {df['percentage'].min():.2f}%")
    
    if 'grade' in df.columns:
        print(f"\n🏆 Grade Distribution:")
        grade_counts = df['grade'].value_counts().sort_index()
        for grade, count in grade_counts.items():
            pct = (count / total_students) * 100
            print(f"   {grade}: {count} students ({pct:.1f}%)")
    
    if 'examName' in df.columns:
        print(f"\n📚 Exams:")
        exam_counts = df['examName'].value_counts()
        for exam, count in exam_counts.items():
            print(f"   {exam}: {count} submissions")
    
    print("=" * 50)


def main():
    print("🔥 Firestore to Excel Exporter")
    print("-" * 40)
    
    db = initialize_firebase()
    if not db:
        return
    
    print("📡 Fetching data from Firestore...")
    data = fetch_quizResults(db)
    
    if not data:
        print("⚠️  No documents found in 'quizResults' collection")
        print("   Make sure students have completed quizzes!")
        return
    
    print(f"✅ Found {len(data)} documents")
    
    print_statistics(data)
    
    print("\n📤 Exporting to Excel...")
    output_path = export_to_excel(data)
    
    print(f"✅ Excel file saved: {output_path}")
    print("\n💡 Tip: Check the 'exports' folder for your files!")


if __name__ == '__main__':
    main()
